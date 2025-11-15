import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait as W
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl chưa được cài đặt. Chạy: pip install openpyxl")

BASE_URL = os.environ.get("MEDSTORE_BASE_URL", "http://localhost:5173")
LOGIN_URL = os.environ.get("MEDSTORE_LOGIN_URL", f"{BASE_URL}/login")

HEADLESS = os.environ.get("HEADLESS", "1").lower() not in ("0", "false", "no")
SLOW_MS = int(os.environ.get("SLOW_MS", "500"))
WAIT_TIMEOUT = int(os.environ.get("WAIT_TIMEOUT", "25"))

SEARCH_EXCEL = r"D:\Downloads\Medical-Store-frontend\Medical-Store-frontend\frontend\KTPM_PHU\DataTest\Search.xlsx"
QUESTION_EXCEL = r"D:\Downloads\Medical-Store-frontend\Medical-Store-frontend\frontend\KTPM_PHU\DataTest\Question.xlsx"
CART_EXCEL = r"D:\Downloads\Medical-Store-frontend\Medical-Store-frontend\frontend\KTPM_PHU\DataTest\Cart.xlsx"
RESULT_EXCEL = r"D:\Downloads\Medical-Store-frontend\Medical-Store-frontend\frontend\KTPM_PHU\Result\ResultSearchQuestionCart.xlsx"


def slow(multiplier: float = 1.0):
    if SLOW_MS > 0:
        time.sleep(SLOW_MS / 1000.0 * multiplier)


def create_driver():
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1440,960")
    try:
        return webdriver.Chrome(options=opts)
    except Exception:
        return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)


def first_present(drv, xpaths, timeout=WAIT_TIMEOUT):
    last_err = None
    for xp in xpaths:
        try:
            return W(drv, timeout).until(
                EC.presence_of_element_located((By.XPATH, xp))
            )
        except Exception as e:
            last_err = e
    raise last_err or TimeoutException("element not found")


def first_clickable(drv, xpaths, timeout=WAIT_TIMEOUT):
    last_err = None
    for xp in xpaths:
        try:
            return W(drv, timeout).until(
                EC.element_to_be_clickable((By.XPATH, xp))
            )
        except Exception as e:
            last_err = e
    raise last_err or TimeoutException("element not clickable")


def js_click(drv, el):
    drv.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    drv.execute_script("arguments[0].click();", el)


def open_home(d):
    d.get(BASE_URL)
    slow(0.7)


def open_login(d):
    d.get(LOGIN_URL)
    slow(0.7)
    first_present(
        d,
        [
            "//h3[contains(normalize-space(.),'Chào mừng')]",
            "//h4[contains(normalize-space(.),'Chào mừng')]",
            "//button[contains(normalize-space(.),'Đăng nhập')]",
        ],
        timeout=WAIT_TIMEOUT,
    )


def field_by_label_or_name(drv, label_text=None, name_candidates=()):
    c = []
    if label_text:
        c += [
            f"//label[contains(normalize-space(.), '{label_text}')]/following::*[self::input or self::textarea][1]",
            f"//*[contains(normalize-space(.), '{label_text}')]/following::*[self::input or self::textarea][1]",
        ]
    for nm in name_candidates:
        c.append(f"//input[@name='{nm}']")
    return first_present(drv, c, timeout=WAIT_TIMEOUT)


def button_by_text(drv, text, timeout=WAIT_TIMEOUT):
    return first_clickable(
        drv,
        [
            f"//button[normalize-space()='{text}']",
            f"//button[contains(normalize-space(.), '{text}')]",
        ],
        timeout=timeout,
    )


def login_user(d, username, password):
    open_login(d)
    u = field_by_label_or_name(d, label_text="Tên đăng nhập", name_candidates=("username",))
    p = field_by_label_or_name(d, label_text="Mật khẩu", name_candidates=("password",))

    try:
        u.clear()
    except Exception:
        pass
    u.send_keys(username)
    slow(0.2)

    try:
        p.clear()
    except Exception:
        pass
    p.send_keys(password)
    slow(0.2)

    p.send_keys(Keys.ENTER)
    slow(0.5)

    W(d, WAIT_TIMEOUT).until(lambda drv: "/login" not in drv.current_url)
    slow(0.5)


def search_keyword(d, keyword):
    open_home(d)

    search_input = first_present(
        d,
        [
            "//input[contains(@placeholder,'Tìm kiếm thuốc')]",
            "//input[contains(@placeholder,'Tìm kiếm')]",
            "//input[@type='search']",
            "(//header//input)[1]",
        ],
        timeout=WAIT_TIMEOUT,
    )
    try:
        search_input.clear()
    except Exception:
        pass

    search_input.send_keys(keyword)
    slow(0.3)

    search_input.send_keys(Keys.ENTER)
    slow(0.5)

    first_present(
        d,
        [
            "//*[contains(normalize-space(.),'Kết quả tìm kiếm cho')]",
            "//*[contains(normalize-space(.),'Không tìm thấy sản phẩm nào')]",
        ],
        timeout=WAIT_TIMEOUT,
    )


def assert_no_results(d):
    el = first_present(
        d,
        ["//*[contains(normalize-space(.),'Không tìm thấy sản phẩm nào')]"],
        timeout=WAIT_TIMEOUT,
    )
    assert "Không tìm thấy sản phẩm nào" in el.text


def open_first_result_and_get_name(d):
    name_el = first_present(
        d,
        [
            "(//div[contains(@class,'MuiCard-root')]//h3)[1]",
            "(//div[contains(@class,'MuiCard-root')]//h6)[1]",
        ],
        timeout=WAIT_TIMEOUT,
    )
    product_name = name_el.text.strip()

    card_el = name_el.find_element(
        By.XPATH, "./ancestor::div[contains(@class,'MuiCard-root')]"
    )
    js_click(d, card_el)
    slow(0.7)

    first_present(
        d,
        [
            f"//*[contains(normalize-space(.),\"{product_name}\")]",
            "//h4[contains(normalize-space(.),'Chi tiết sản phẩm')]",
        ],
        timeout=WAIT_TIMEOUT,
    )

    return product_name


def go_to_question_box(d):
    textarea = first_present(
        d,
        [
            "//textarea[contains(@placeholder,'Nhập câu hỏi của bạn về sản phẩm')]",
            "//textarea",
        ],
        timeout=WAIT_TIMEOUT,
    )
    d.execute_script("arguments[0].scrollIntoView({block:'center'});", textarea)
    slow(0.5)
    return textarea


def go_to_cart_page_via_header(d):
    cart_label = first_present(
        d,
        [
            "//header//*[normalize-space()='Giỏ hàng']",
            "//*[contains(@class,'MuiAppBar-root')]//*[normalize-space()='Giỏ hàng']",
            "//*[normalize-space()='Giỏ hàng']",
        ],
        timeout=WAIT_TIMEOUT,
    )
    js_click(d, cart_label)
    slow(0.7)

    first_present(
        d,
        ["//*[contains(normalize-space(.),'Giỏ hàng của tôi')]"],
        timeout=WAIT_TIMEOUT,
    )


def test_search_with_data_from_excel():
    search_data = load_search_data_from_excel()
    results = []
    
    for idx, test_case in enumerate(search_data, 1):
        keyword = test_case['keyword']
        expected_result = test_case['expected_result']
        
        pass_fail = "FAIL"
        
        d = create_driver()
        try:
            search_keyword(d, keyword)
            
            if expected_result.upper() == "NO_RESULTS":
                assert_no_results(d)
                pass_fail = "PASS"
                print(f"✅ Test case {idx} PASS: Tìm '{keyword}' không ra sản phẩm nào")
            elif expected_result.upper() == "HAS_RESULTS":
                first_present(
                    d,
                    ["(//div[contains(@class,'MuiCard-root')])[1]"],
                    timeout=WAIT_TIMEOUT,
                )
                pass_fail = "PASS"
                print(f"✅ Test case {idx} PASS: Tìm '{keyword}' có sản phẩm")
        except Exception as e:
            pass_fail = "FAIL"
            print(f"❌ Test case {idx} FAIL: {e}")
        finally:
            d.quit()
        
        results.append({
            'test_case': f"Search-{idx}",
            'username': '',
            'password': '',
            'expected_result': expected_result,
            'pass_fail': pass_fail
        })
    
    return results


def test_question_with_data_from_excel():
    question_data = load_question_data_from_excel()
    results = []
    
    for idx, test_case in enumerate(question_data, 1):
        precondition = test_case['precondition']
        keyword = test_case['keyword']
        username = test_case['username']
        password = test_case['password']
        question = test_case['question']
        expected_result = test_case['expected_result']
        
        pass_fail = "FAIL"
        
        d = create_driver()
        try:
            if precondition.lower() == "đã đăng nhập" and username and password:
                login_user(d, username, password)
            
            search_keyword(d, keyword)
            open_first_result_and_get_name(d)
            
            textarea = go_to_question_box(d)
            textarea.send_keys(question)
            slow(0.3)
            
            ask_btn = button_by_text(d, "Đặt câu hỏi")
            js_click(d, ask_btn)
            
            if "đăng nhập" in expected_result.lower():
                msg = first_present(
                    d,
                    [
                        "//*[contains(normalize-space(.),'Vui lòng đăng nhập trước khi đặt câu hỏi về sản phẩm')]",
                        "//*[contains(normalize-space(.),'Yêu cầu đăng nhập')]",
                    ],
                    timeout=WAIT_TIMEOUT,
                )
                assert "đăng nhập" in msg.text.lower()
                assert expected_result.lower() in msg.text.lower() or msg.text.lower() in expected_result.lower()
                pass_fail = "PASS"
                print(f"✅ Test case {idx} PASS: {expected_result}")
            elif "thành công" in expected_result.lower():
                snackbar = first_present(
                    d,
                    ["//*[contains(normalize-space(.),'Đã đặt câu hỏi thành công')]"],
                    timeout=WAIT_TIMEOUT,
                )
                assert "Đã đặt câu hỏi thành công" in snackbar.text
                assert expected_result.lower() in snackbar.text.lower() or snackbar.text.lower() in expected_result.lower()
                pass_fail = "PASS"
                print(f"✅ Test case {idx} PASS: {expected_result}")
        except Exception as e:
            pass_fail = "FAIL"
            print(f"❌ Test case {idx} FAIL: {e}")
        finally:
            d.quit()
        
        results.append({
            'test_case': f"Question-{idx}",
            'username': username,
            'password': password,
            'expected_result': expected_result,
            'pass_fail': pass_fail
        })
    
    return results


def assert_error_alert(d):
    al = first_present(d, ["//*[@role='alert']"], timeout=WAIT_TIMEOUT)
    txt = (al.text or "").lower()
    assert "đăng nhập" in txt or "lỗi" in txt or "thất bại" in txt or "error" in txt


def load_search_data_from_excel():
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl chưa được cài đặt. Chạy: pip install openpyxl")
    
    if not os.path.exists(SEARCH_EXCEL):
        raise FileNotFoundError(f"File Excel không tồn tại: {SEARCH_EXCEL}")
    
    try:
        wb = load_workbook(SEARCH_EXCEL)
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            keyword = str(row[0]).strip() if row[0] else ''
            expected_result = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            
            if not keyword:
                continue
            
            data.append({
                'keyword': keyword,
                'expected_result': expected_result
            })
        
        wb.close()
        print(f"✅ Đã đọc {len(data)} bản ghi từ {SEARCH_EXCEL}")
        return data
    except Exception as e:
        raise Exception(f"Lỗi khi đọc Excel Search: {e}")


def load_question_data_from_excel():
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl chưa được cài đặt. Chạy: pip install openpyxl")
    
    if not os.path.exists(QUESTION_EXCEL):
        raise FileNotFoundError(f"File Excel không tồn tại: {QUESTION_EXCEL}")
    
    try:
        wb = load_workbook(QUESTION_EXCEL)
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            precondition = str(row[0]).strip() if row[0] else ''
            keyword = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            username = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            password = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            question = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            expected_result = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            
            if not keyword:
                continue
            
            data.append({
                'precondition': precondition,
                'keyword': keyword,
                'username': username,
                'password': password,
                'question': question,
                'expected_result': expected_result
            })
        
        wb.close()
        print(f"✅ Đã đọc {len(data)} bản ghi từ {QUESTION_EXCEL}")
        return data
    except Exception as e:
        raise Exception(f"Lỗi khi đọc Excel Question: {e}")


def load_cart_data_from_excel():
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl chưa được cài đặt. Chạy: pip install openpyxl")
    
    if not os.path.exists(CART_EXCEL):
        raise FileNotFoundError(f"File Excel không tồn tại: {CART_EXCEL}")
    
    try:
        wb = load_workbook(CART_EXCEL)
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            precondition = str(row[0]).strip() if row[0] else ''
            keyword = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            username = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            password = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            expected_result = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            
            if not keyword:
                continue
            
            data.append({
                'precondition': precondition,
                'keyword': keyword,
                'username': username,
                'password': password,
                'expected_result': expected_result
            })
        
        wb.close()
        print(f"✅ Đã đọc {len(data)} bản ghi từ {CART_EXCEL}")
        return data
    except Exception as e:
        raise Exception(f"Lỗi khi đọc Excel Cart: {e}")


def export_results_to_excel(results):
    if not OPENPYXL_AVAILABLE:
        print("⚠️ Không thể xuất kết quả ra Excel vì thiếu thư viện openpyxl")
        return
    
    os.makedirs(os.path.dirname(RESULT_EXCEL), exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    headers = ["Test Case", "Username", "Password", "Expected Result", "Pass/Fail"]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_style = Side(style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=result['test_case']).border = border
        ws.cell(row=row_idx, column=2, value=result.get('username', '')).border = border
        ws.cell(row=row_idx, column=3, value=result.get('password', '')).border = border
        ws.cell(row=row_idx, column=4, value=result['expected_result']).border = border
        
        pass_fail_cell = ws.cell(row=row_idx, column=5, value=result['pass_fail'])
        pass_fail_cell.border = border
        pass_fail_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if result['pass_fail'].upper() == "PASS":
            pass_fail_cell.fill = pass_fill
            pass_fail_cell.font = Font(bold=True, color="006100")
        else:
            pass_fail_cell.fill = fail_fill
            pass_fail_cell.font = Font(bold=True, color="9C0006")
    
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25
    
    ws.row_dimensions[1].height = 30
    
    for row_idx in range(2, len(results) + 2):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx != 5:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    wb.save(RESULT_EXCEL)
    print(f"✅ Đã xuất kết quả test vào {RESULT_EXCEL}")


def test_cart_with_data_from_excel():
    cart_data = load_cart_data_from_excel()
    results = []
    
    for idx, test_case in enumerate(cart_data, 1):
        precondition = test_case['precondition']
        keyword = test_case['keyword']
        username = test_case['username']
        password = test_case['password']
        expected_result = test_case['expected_result']
        
        pass_fail = "FAIL"
        
        d = create_driver()
        try:
            if precondition.lower() == "đã đăng nhập" and username and password:
                login_user(d, username, password)
            
            search_keyword(d, keyword)
            product_name = open_first_result_and_get_name(d)
            
            add_btn = button_by_text(d, "Thêm vào giỏ")
            js_click(d, add_btn)
            
            if "đăng nhập" in expected_result.lower():
                try:
                    assert_error_alert(d)
                except TimeoutException:
                    msg = first_present(
                        d,
                        [
                            "//*[contains(normalize-space(.),'Vui lòng đăng nhập trước khi mua hoặc thêm sản phẩm vào giỏ hàng')]",
                            "//*[contains(normalize-space(.),'Yêu cầu đăng nhập')]",
                        ],
                        timeout=WAIT_TIMEOUT,
                    )
                    assert "đăng nhập" in msg.text.lower()
                    assert expected_result.lower() in msg.text.lower() or msg.text.lower() in expected_result.lower()
                pass_fail = "PASS"
                print(f"✅ Test case {idx} PASS: {expected_result}")
            elif keyword.lower() in expected_result.lower():
                try:
                    first_present(
                        d,
                        ["//*[contains(normalize-space(.),'Đã thêm sản phẩm vào giỏ hàng')]"],
                        timeout=WAIT_TIMEOUT,
                    )
                except TimeoutException:
                    pass
                
                go_to_cart_page_via_header(d)
                
                item = first_present(
                    d,
                    [
                        f"//*[contains(normalize-space(.),\"{product_name}\")]",
                    ],
                    timeout=WAIT_TIMEOUT,
                )
                assert product_name.strip() in item.text
                assert keyword.lower() in item.text.lower() or product_name.lower() in expected_result.lower()
                pass_fail = "PASS"
                print(f"✅ Test case {idx} PASS: {expected_result}")
        except Exception as e:
            pass_fail = "FAIL"
            print(f"❌ Test case {idx} FAIL: {e}")
        finally:
            d.quit()
        
        results.append({
            'test_case': f"Cart-{idx}",
            'username': username,
            'password': password,
            'expected_result': expected_result,
            'pass_fail': pass_fail
        })
    
    return results


if __name__ == "__main__":
    print(f"HEADLESS={HEADLESS} SLOW_MS={SLOW_MS} BASE={BASE_URL}")
    
    all_results = []
    
    try:
        print("\n📊 Chạy test Search...")
        search_results = test_search_with_data_from_excel()
        all_results.extend(search_results)
        
        print("\n📊 Chạy test Question...")
        question_results = test_question_with_data_from_excel()
        all_results.extend(question_results)
        
        print("\n📊 Chạy test Cart...")
        cart_results = test_cart_with_data_from_excel()
        all_results.extend(cart_results)
        
        print("\n📊 Xuất kết quả ra Excel...")
        export_results_to_excel(all_results)
        
    except FileNotFoundError as e:
        print(f"\n❌ Lỗi: {e}")
        print("   Vui lòng đảm bảo các file Excel tồn tại tại đường dẫn đã chỉ định.")
    except ImportError as e:
        print(f"\n❌ Lỗi: {e}")
    except Exception as e:
        print(f"\n❌ Lỗi không mong đợi: {e}")
        raise
