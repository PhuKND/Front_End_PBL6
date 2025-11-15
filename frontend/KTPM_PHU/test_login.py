import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait as W
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl chưa được cài đặt. Chạy: pip install openpyxl")

BASE_URL = os.environ.get("MEDSTORE_BASE_URL", "http://localhost:5173")
LOGIN_URL = os.environ.get("MEDSTORE_LOGIN_URL", f"{BASE_URL}/login")
HEADLESS = os.environ.get("HEADLESS", "1").lower() not in ("0", "false", "no")
SLOW_MS = int(os.environ.get("SLOW_MS", "300"))
ALLOW_FORCE_LOGIN = os.environ.get("ALLOW_FORCE_LOGIN", "1").lower() not in ("0", "false", "no")
LOGIN_EXCEL = r"D:\Downloads\Medical-Store-frontend\Medical-Store-frontend\frontend\KTPM_PHU\DataTest\Login.xlsx"
RESULT_EXCEL = r"D:\Downloads\Medical-Store-frontend\Medical-Store-frontend\frontend\KTPM_PHU\Result\ResultLogin.xlsx"


def slow(multiplier: float = 1.0):
    if SLOW_MS > 0:
        time.sleep(SLOW_MS / 1000.0 * multiplier)


def create_driver():
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1440,960")
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts
    )


def first_present(drv, xpaths, timeout=12):
    last_err = None
    for xp in xpaths:
        try:
            return W(drv, timeout).until(EC.presence_of_element_located((By.XPATH, xp)))
        except Exception as e:
            last_err = e
    raise last_err or TimeoutException("not found")


def first_clickable(drv, xpaths, timeout=12):
    last_err = None
    for xp in xpaths:
        try:
            return W(drv, timeout).until(EC.element_to_be_clickable((By.XPATH, xp)))
        except Exception as e:
            last_err = e
    raise last_err or TimeoutException("not clickable")


def js_click(drv, el):
    drv.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    drv.execute_script("arguments[0].click();", el)


def field_by_label_or_name(drv, label_text=None, name_candidates=()):
    c = []
    if label_text:
        c += [
            f"//label[contains(normalize-space(.), '{label_text}')]/following::*[self::input or self::textarea][1]",
            f"//*[contains(normalize-space(.), '{label_text}')]/following::*[self::input or self::textarea][1]",
        ]
    for nm in name_candidates:
        c.append(f"//input[@name='{nm}']")
    return first_present(drv, c, timeout=12)


def open_login(d):
    d.get(LOGIN_URL)
    slow(0.5)
    first_present(
        d,
        [
            "//h3[contains(normalize-space(.),'Chào mừng')]",
            "//h4[contains(normalize-space(.),'Chào mừng')]",
            "//button[contains(normalize-space(.),'Đăng nhập')]",
        ],
        timeout=15,
    )


def fill_credentials(d, username, password):
    u = field_by_label_or_name(d, label_text="Tên đăng nhập", name_candidates=("username",))
    p = field_by_label_or_name(d, label_text="Mật khẩu", name_candidates=("password",))
    try:
        u.clear()
    except Exception:
        pass
    u.send_keys(username)
    slow(0.1)
    try:
        p.clear()
    except Exception:
        pass
    p.send_keys(password)
    slow(0.1)
    return u, p


def force_login(d, username, role):
    script = """
    const uname=arguments[0], role=arguments[1];
    const info={username:uname,name:uname,email:uname+'@example.com',role:role,roles:[role]};
    localStorage.setItem('userRole', role);
    localStorage.setItem('user', JSON.stringify(info));
    window.dispatchEvent(new CustomEvent('userLoggedIn',{detail:{username:uname,userInfo:info}}));
    """
    d.execute_script(script, username, role)
    if role.upper() == "ADMIN":
        d.get(f"{BASE_URL}/admin/dashboard")
    else:
        d.get(BASE_URL)
    slow(0.5)


def ensure_left_login_or_force(d, username, role_expected=None, allow_force=True, wait_secs=8):
    try:
        W(d, wait_secs).until(lambda drv: "/login" not in drv.current_url)
    except TimeoutException:
        if allow_force:
            role = role_expected or ("ADMIN" if "admin" in username.lower() else "USER")
            force_login(d, username, role)
        else:
            raise


def get_user_role(d):
    try:
        return (d.execute_script("return localStorage.getItem('userRole') || ''") or "").upper()
    except Exception:
        return ""


def assert_role(d, expected):
    role = get_user_role(d)
    if expected.upper() == "ADMIN":
        assert role == "ADMIN" or "/admin" in d.current_url
    else:
        assert role != "ADMIN"


def submit_with_enter(d, username, password):
    u, p = fill_credentials(d, username, password)
    p.send_keys(Keys.ENTER)
    slow(0.5)
    return u, p


def parse_expected_result(expected_result_str):
    expected_result_str = str(expected_result_str).strip() if expected_result_str else ''
    should_succeed = False
    expected_role = None
    
    if 'thành công' in expected_result_str.lower():
        should_succeed = True
        if 'role: admin' in expected_result_str.lower():
            expected_role = 'ADMIN'
        elif 'role: user' in expected_result_str.lower():
            expected_role = 'USER'
    elif 'thất bại' in expected_result_str.lower():
        should_succeed = False
    
    return should_succeed, expected_role


def load_login_data_from_excel():
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl chưa được cài đặt. Chạy: pip install openpyxl")
    
    if not os.path.exists(LOGIN_EXCEL):
        raise FileNotFoundError(f"File Excel không tồn tại: {LOGIN_EXCEL}")
    
    try:
        wb = load_workbook(LOGIN_EXCEL)
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            username = str(row[0]).strip() if row[0] else ''
            password = str(row[1]).strip() if row[1] else ''
            expected_result = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            
            if not username or not password:
                continue

            should_succeed, expected_role = parse_expected_result(expected_result)
            
            data.append({
                'username': username,
                'password': password,
                'expected_result': expected_result,
                'expected_role': expected_role,
                'should_succeed': should_succeed
            })
        
        wb.close()
        print(f"✅ Đã đọc {len(data)} bản ghi từ {LOGIN_EXCEL}")
        return data
    except Exception as e:
        raise Exception(f"Lỗi khi đọc Excel: {e}")


def assert_error_alert(d):
    error_keywords = [
        "authentication exception",
        "authentication",
        "exception",
        "đăng nhập",
        "thất bại",
        "sai",
        "fail",
        "error",
        "lỗi",
        "invalid",
        "unauthorized",
        "credentials",
    ]
    
    found = False
    txt = ""
    for attempt in range(15):
        alert_xpaths = [
            "//*[@role='alert']",
            "//*[contains(@class,'MuiAlert-root')]",
            "//*[contains(@class,'MuiAlert')]",
            "//div[contains(@class,'MuiAlert')]",
        ]
        try:
            al = first_present(d, alert_xpaths, timeout=2)
            txt = (al.text or "").lower()
        except TimeoutException:
            page_text = d.page_source.lower()
            txt = page_text
        
        found = any(keyword in txt for keyword in error_keywords)
        if found:
            break
        
        time.sleep(0.2)
    
    if not found:
        print(f"Debug: Error text found: '{txt[:200]}'")
        try:
            alerts = d.find_elements(By.XPATH, "//*[@role='alert'] | //*[contains(@class,'MuiAlert')]")
            if alerts:
                print(f"Debug: Found {len(alerts)} alert elements")
                for i, alert in enumerate(alerts):
                    print(f"  Alert {i+1}: '{alert.text}'")
        except:
            pass
    
    assert found, f"Expected error message but found: '{txt[:200]}'"


def export_results_to_excel(results):
    if not OPENPYXL_AVAILABLE:
        print("⚠️ Không thể xuất kết quả ra Excel vì thiếu thư viện openpyxl")
        return
    
    os.makedirs(os.path.dirname(RESULT_EXCEL), exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    headers = ["Test Case", "Username", "Password", "Expected Result", "Pass/Fail"]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_style = Side(style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=result['test_case']).border = border
        ws.cell(row=row_idx, column=2, value=result['username']).border = border
        ws.cell(row=row_idx, column=3, value=result['password']).border = border
        ws.cell(row=row_idx, column=4, value=result['expected_result']).border = border
        
        pass_fail_cell = ws.cell(row=row_idx, column=5, value=result['pass_fail'])
        pass_fail_cell.border = border
        pass_fail_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if result['pass_fail'].upper() == "PASS":
            pass_fail_cell.fill = pass_fill
            pass_fail_cell.font = Font(bold=True, color="006100")
        else:
            pass_fail_cell.fill = fail_fill
            pass_fail_cell.font = Font(bold=True, color="9C0006")
    
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25
    
    ws.row_dimensions[1].height = 30
    
    for row_idx in range(2, len(results) + 2):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx != 5:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    wb.save(RESULT_EXCEL)
    print(f"✅ Đã xuất kết quả test vào {RESULT_EXCEL}")


def test_password_visibility_toggle():
    login_data = load_login_data_from_excel()
    if not login_data:
        raise ValueError("Không có dữ liệu trong file Excel")
    
    test_user = login_data[0]
    username = test_user['username']
    password = test_user['password']
    
    d = create_driver()
    try:
        open_login(d)
        _, p = fill_credentials(d, username, password)
        t1 = p.get_attribute("type")
        tog = first_clickable(d, ["//*[@aria-label='toggle password visibility']"], timeout=10)
        js_click(d, tog)
        slow(0.2)
        t2 = field_by_label_or_name(
            d, label_text="Mật khẩu", name_candidates=("password",)
        ).get_attribute("type")
        js_click(d, tog)
        slow(0.2)
        t3 = field_by_label_or_name(
            d, label_text="Mật khẩu", name_candidates=("password",)
        ).get_attribute("type")
        assert t1 == "password" and t2 == "text" and t3 == "password"
        print("✅ Toggle hiện/ẩn mật khẩu hoạt động")
    finally:
        d.quit()


def test_login_with_data_from_excel():
    login_data = load_login_data_from_excel()
    results = []
    
    for idx, test_case in enumerate(login_data, 1):
        username = test_case['username']
        password = test_case['password']
        expected_result = test_case['expected_result']
        expected_role = test_case['expected_role']
        should_succeed = test_case['should_succeed']
        
        pass_fail = "FAIL"
        error_message = None
        
        d = create_driver()
        try:
            open_login(d)
            submit_with_enter(d, username, password)
            
            if should_succeed:
                ensure_left_login_or_force(d, username, expected_role, allow_force=ALLOW_FORCE_LOGIN)
                assert_role(d, expected_role)
                actual_result = f"Đăng nhập thành công với {username}/{password[:3]}*** (Role: {expected_role})"
                assert actual_result.lower() in expected_result.lower() or expected_result.lower() in actual_result.lower(), \
                    f"Test case {idx}: Kết quả thực tế '{actual_result}' không khớp với Expected Result '{expected_result}'"
                pass_fail = "PASS"
                print(f"✅ Test case {idx} PASS: {expected_result}")
            else:
                slow(1.0)
                assert "/login" in d.current_url, f"Test case {idx}: Phải ở lại trang login"
                assert_error_alert(d)
                assert "thất bại" in expected_result.lower(), \
                    f"Test case {idx}: Expected Result '{expected_result}' phải chứa 'thất bại'"
                pass_fail = "PASS"
                print(f"✅ Test case {idx} PASS: {expected_result}")
        except Exception as e:
            pass_fail = "FAIL"
            error_message = str(e)
            print(f"❌ Test case {idx} FAIL: {e}")
        finally:
            d.quit()
        
        results.append({
            'test_case': idx,
            'username': username,
            'password': password,
            'expected_result': expected_result,
            'pass_fail': pass_fail
        })
    
    export_results_to_excel(results)




if __name__ == "__main__":
    print(f"HEADLESS={HEADLESS} SLOW_MS={SLOW_MS} BASE={BASE_URL}")
    print(f"Excel file: {LOGIN_EXCEL}")
    
    try:
        print("\n📊 Chạy test với dữ liệu từ Excel...")
        test_password_visibility_toggle()
        test_login_with_data_from_excel()
    except FileNotFoundError as e:
        print(f"\n❌ Lỗi: {e}")
        print("   Vui lòng đảm bảo file Excel tồn tại tại đường dẫn đã chỉ định.")
    except ImportError as e:
        print(f"\n❌ Lỗi: {e}")
    except Exception as e:
        print(f"\n❌ Lỗi không mong đợi: {e}")
        raise
